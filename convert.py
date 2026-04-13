import time
from openpyxl.styles import Alignment
from openpyxl import Workbook, load_workbook
import requests
import logging

logging.basicConfig(filename="newfile.log",
                    format='%(asctime)s %(levelname)s: %(message)s',
                    filemode='w')

col1 = "NCT"
col2 = "URL"
col3 = "Intervention/Treatment"
col4 = "Trial Name"
col5 = "Acronym"
col6 = "Sponsor"
col7 = "Funder Type"
col8 = "Phase"
col9 = "Enrollment"
col10 = "Status"
col11 = "Brief Summary"
col12 = "Detailed Description"
col13 = "Inclusion Criteria (may also contain the Exclusion Criteria due to parsing issues)"
col14 = "Exclusion Criteria"
col15 = "Conditions"
col16 = "Primary Purpose"
col17 = "Primary Outcome Measures"
col18 = "Secondary Outcome Measures"
col19 = "Allocation"
col20 = "Interventional Model"
col21 = "Masking"
col22 = "Study Start"
col23 = "Primary Completion Date (Est)"
col24 = "Locations"

#A naive and hacky way of finding what we want in a json. 
#While structure names in the json's are almost always consistent, actual structure isnt.
#so we just recusively search through the entire json until we find the first thing matching our paramter.
#this of course means you better make sure whatever you're searching for doesn't share a name with anything else.
#If what you're rtrying to find does share a name, you should instead try to search for the closest uniquely named parent and then
#manually access each dictionary key until you get what you need   
def find_key_recursive(data, target_key):
    if isinstance(data, dict):
        for key, value in data.items():
            if key == target_key:
                return value
            result = find_key_recursive(value, target_key)
            if result is not None:
                return result
    elif isinstance(data, list):
        for item in data:
            result = find_key_recursive(item, target_key)
            if result is not None:
                return result
    return None

#We only want certain parts of the location data, a lot of it is TMI
def locParse(dict):
    str2 = ""
    for key, value in dict.items():
        str = ""
        if (key != "facility" and key != "city" and key != "state" and key != "country" and key != "zip"):
            continue
        str += f", {value}"
        str2 += str
        if (len(str2)) > 1000:
            str2 += ", [TRUNCATED to 1000 characters for processing reasons]"
            break
    return str2[2:]

#Most* NCT's have their eligibility clearly laid out with the words "Inclusion Criteria" and "Exclusion Criteria"
#We can generally split our criteria page using this info to get the critera's serpated
#SOME assholes like to be "different" and don't adhere to this unwritten standard, so rarely this fails to seperate the two
def splitCriteria(str):
    ind = str.find("Exclusion Criteria")
    if ind != -1:
        first_part = str[:ind]
        second_part = str[ind:]
        return [first_part, second_part]
    else:
        return [str]

ncts = ""              #1
urls = ""              #2
interventions = ""     #3
names = ""             #4
acronyms = ""          #5
sponsors = ""          #6
funderTypes = ""       #7
phases = ""            #8
enrolls = ""           #9
statuses = ""           #10
summaries = ""          #11
descriptions = ""       #12
criterias = ""          #13
exclusions = ""         #14
conditions = ""         #15
primPurposes = ""       #16
primOutcomes = ""       #17
secOutcomes = ""        #18
allocations = ""        #19
interventionModels = ""  #20
maskings = ""            #21
starts = ""              #22
primComps = ""           #23
locs = ""                #24

#Had issues where randomly a request would hang indefinitely and never recover.
#really messed up proccessing when the amount of NCT's we have to process is in the thousands.
#This supposedly fixes the issue by redialling, but its hard to test for something that is completely random
#so I can't say for sure that it fixes the issue.
def httpreq(url, timeout=5, max_retries=3, backoff=1):
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.get(url, timeout=timeout)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.Timeout:
            logger.warning(f"Timeout on attempt {attempt}/{max_retries} for URL: {url}")
            if attempt == max_retries:
                logger.error(f"Max retries reached for URL: {url}")
                return None
            time.sleep(backoff)
            backoff *= 2
        except requests.exceptions.RequestException as e:
            logger.error(f"Request failed for URL: {url} - {e}")
            return None

    return None 

#the meat and potatoes of the program, this is where we get all the data for each NCT and fill the respective row with it.
def mainLoop(nctNum, ws, row_num, logger):
    url: str = f'http://clinicaltrials.gov/api/v2/studies/{nctNum}'

    if (row_num % 200 == 0):
        logger.info(f"Processing NCT {nctNum} at row {row_num}...")

    data = httpreq(url)
    if data is None:
        logger.warning(f"Skipping NCT {nctNum} after request failure.")
        return

        
    det_desc = find_key_recursive(data, "detailedDescription")
    crit = find_key_recursive(data, "eligibilityCriteria")
    
    #This should basically NEVER happen, but I need to make sure nothing goes wrong if it somehow does
    if (crit == None):
        logger.warning(f"Eligibility criteria not found for NCT {nctNum}. Weird.") 
        critList = None
    else:
        critList = splitCriteria(crit)
    name = find_key_recursive(data, "briefTitle")
    nct = find_key_recursive(data, "nctId")
    sponsor = find_key_recursive(data, "name")
    studyStart = find_key_recursive(data, "startDateStruct") 
    primComp = find_key_recursive(data, "primaryCompletionDateStruct") 
    enrollment = find_key_recursive(data, "enrollmentInfo") 
    phase = find_key_recursive(data, "phases") 
    brief_summary = find_key_recursive(data, "briefSummary")
    Intervention_treat = find_key_recursive(data, "interventionNames") 
    status = find_key_recursive(data, "overallStatus")
    designDetails = find_key_recursive(data, "designInfo") 
    loc = find_key_recursive(data, "locations")
    acronym = find_key_recursive(data, "acronym")
    condition = find_key_recursive(data, "conditions")
    funderType = find_key_recursive(data, "organization")["class"]
    outcomes = find_key_recursive(data, "outcomesModule")

    '''
    Now comes the very ugly part. Ignoring the first couple of vars, which should be present no matter what
    (unless something goes very very wrong with either your input excel sheet or the API) we need to check 
    if each var is None before we can use it, because some NCT's either don't have the info we looked for or its somehow named
    something completely different from every other NCT (I don't have the energy or time to account for every fucking naming possibility). 
    Thankfully I can count on one hand the amount of NCT's that name things very differently
    so this is mostly just to make sure nothing breaks when we run into those few outliers. 

    So we have a shit-ton of if statements. If the var is None, we make the string an empty string. Otherwise we just
    fill in the data accordingly.
    '''

    ncts = nct
    
    urls = f"https://clinicaltrials.gov/study/{nctNum}"

    if name == None:
        names = ""
    else:
        names = name
    
    if phase == None:
        phases = ""
    else:
        phases = (' '.join(str(x) for x in phase))
    
    if sponsor == None:
        sponsors = ""
    else:
        sponsors = sponsor
    
    if status == None:
        statuses = ""
    else:
        statuses = status
    
    if acronym == None:
        acronyms = ""
    else:
        acronyms = acronym

    if brief_summary == None:
        summaries = ""
    else:
        summaries = brief_summary
    
    if det_desc == None:
        descriptions = ""
    else:
        descriptions = det_desc
    
    #again, this should never happen, but just in case something goes very wrong with the parsing of the criteria, we want to make sure nothing breaks
    if critList == None:
        criterias = ""
    else:
        criterias = critList[0]
        if (len(critList)) == 2:
            exclusions = critList[1]
        else:
            exclusions = ""

    if (Intervention_treat == None):
        x = find_key_recursive(data, "interventions")
        interventions = ' '.join(str(i) for i in x)
    else:
        interventions = ' '.join(str(x) for x in Intervention_treat)

    if enrollment == None:
        enrolls = ""
    else:
        enrolls = enrollment["count"]
    
    if funderType == None:
        funderTypes = ""
    else:
        funderTypes = funderType

    if condition == None:
        conditions = ""
    else:
        conditions = ', '.join(str(x) for x in condition)[1:]
    
    if outcomes == None:
        primOutcomes = ""
    else:
        primOutcomes = ',\n'.join(str(x["measure"]) for x in outcomes["primaryOutcomes"])
    
    if find_key_recursive(data, "secondaryOutcomes") == None:
        secOutcomes = ""
    else:
        secOutcomes = ',\n'.join(str(x["measure"]) for x in outcomes["secondaryOutcomes"])

    #there was like 1 singular NCT that had no design details. Kind of pissed me off.
    #since those kind of NCT's apparently exist I needed to check for this 
    if designDetails == None:
        primPurposes = ""
        allocations = ""
        interventionModels = ""
        maskings = ""
    else:
        if ("primaryPurpose" not in designDetails):
            primPurposes = ""
        else:
            primPurposes = designDetails["primaryPurpose"]
        
        if ("allocation" not in designDetails):
            allocations = ""
        else:
            allocations = designDetails["allocation"]

        if ("interventionModel" not in designDetails):
            interventionModels = ""
        else:
            interventionModels = designDetails["interventionModel"]

        if ("maskingInfo" not in designDetails):
            maskings = ""
        else:
            maskings = designDetails["maskingInfo"]["masking"]
        

    if (studyStart == None):
        starts = ""
    else:
        starts = studyStart["date"]

    if (primComp == None):
        primComps = ""
    else:
        primComps = primComp["date"]
    
    if (loc == None):
        locs = ""
    else:
        locs = ',\n\n'.join(str(locParse(x)) for x in loc)

    #very ugly way of filling the row with the data we just got. Maybe you could turn this into a loop using ascii values
    #but each string var has a different name so it would be a bit of a pain; simpler but uglier to just do it manually here
    ws[f"A{row_num}"] = ncts
    ws[f"A{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"B{row_num}"] = urls
    ws[f"B{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"C{row_num}"] = interventions
    ws[f"C{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"D{row_num}"] = names
    ws[f"D{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"E{row_num}"] = acronyms
    ws[f"E{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"F{row_num}"] = sponsors
    ws[f"F{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"G{row_num}"] = funderTypes
    ws[f"G{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"H{row_num}"] = phases
    ws[f"H{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"I{row_num}"] = enrolls
    ws[f"I{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"J{row_num}"] = statuses
    ws[f"J{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"K{row_num}"] = summaries
    ws[f"K{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"L{row_num}"] = descriptions
    ws[f"L{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"M{row_num}"] = criterias
    ws[f"M{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"N{row_num}"] = exclusions
    ws[f"N{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"O{row_num}"] = conditions
    ws[f"O{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"P{row_num}"] = primPurposes
    ws[f"P{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"Q{row_num}"] = primOutcomes
    ws[f"Q{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"R{row_num}"] = secOutcomes
    ws[f"R{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"S{row_num}"] = allocations
    ws[f"S{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"T{row_num}"] = interventionModels
    ws[f"T{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"U{row_num}"] = maskings
    ws[f"U{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"V{row_num}"] = starts
    ws[f"V{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"W{row_num}"] = primComps
    ws[f"W{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)
    ws[f"X{row_num}"] = locs
    ws[f"X{row_num}"].alignment = Alignment(vertical='top', wrap_text=True)

    ws.row_dimensions[row_num].height = 80.0

    #reset all string vars, kind of unnecessary but protects against any mistakes where
    #I somehow forgot to fill a string
    ncts = ""              #1
    urls = ""              #2
    interventions = ""     #3
    names = ""             #4
    acronyms = ""          #5
    sponsors = ""          #6
    funderTypes = ""       #7
    phases = ""            #8
    enrolls = ""           #9
    statuses = ""           #10
    summaries = ""          #11
    descriptions = ""       #12
    criterias = ""          #13
    exclusions = ""         #14
    conditions = ""         #15
    primPurposes = ""       #16
    primOutcomes = ""       #17
    secOutcomes = ""        #18
    allocations = ""        #19
    interventionModels = ""  #20
    maskings = ""            #21
    starts = ""              #22
    primComps = ""           #23
    locs = ""                #24
    
    #save every 200 rows just in case, this way if something goes wrong we won't lose all our progress
    if (row_num % 200 == 0):
        wb.save('output.xlsx')
        logger.info("Progress saved.")



if __name__=="__main__":
    #logger stuff
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    #load the workbook and activate
    wb = load_workbook('input.xlsx')
    ws = wb.active

    #get the list of NCT's from the first column, skipping the header
    column_list = [cell.value for cell in ws['A']]
    column_list = column_list[1:]

    print("Working...")

    #create a new workbook for output
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"

    #going row by row, starting at row 2 (i = 2) we run the mainloop which fills the row
    #with the respective NCT's data
    i = 2
    for x in column_list:
        mainLoop(x, ws, i, logger)
        i += 1

    #there might be a more efficient way to do this. Sets column widths what I want them to be.
    ws.column_dimensions['A'].width = 15 #1
    ws.column_dimensions['B'].width = 25 #2
    ws.column_dimensions['C'].width = 25 #3
    ws.column_dimensions['D'].width = 25 #4
    ws.column_dimensions['E'].width = 20 #5
    ws.column_dimensions['F'].width = 25 #6
    ws.column_dimensions['G'].width = 20 #7
    ws.column_dimensions['H'].width = 15 #8
    ws.column_dimensions['I'].width = 15 #9    
    ws.column_dimensions['J'].width = 20 #10
    ws.column_dimensions['K'].width = 50 #11
    ws.column_dimensions['L'].width = 50 #12
    ws.column_dimensions['M'].width = 65 #13
    ws.column_dimensions['N'].width = 50 #14
    ws.column_dimensions['O'].width = 25 #15
    ws.column_dimensions['P'].width = 15 #16
    ws.column_dimensions['Q'].width = 65 #17
    ws.column_dimensions['R'].width = 65 #18
    ws.column_dimensions['S'].width = 25 #19
    ws.column_dimensions['T'].width = 20 #20
    ws.column_dimensions['U'].width = 15 #21
    ws.column_dimensions['V'].width = 15 #22
    ws.column_dimensions['W'].width = 15 #23
    ws.column_dimensions['X'].width = 35 #24

    ws.row_dimensions[1].height = 0.4

    #set header row
    ws["A1"] = "NCT ID"
    ws["B1"] = "URL"
    ws["C1"] = "Intervention/Treatment"
    ws["D1"] = "Trial Name"
    ws["E1"] = "Acronym"
    ws["F1"] = "Sponsor"
    ws["G1"] = "Funder Type"
    ws["H1"] = "Phase"
    ws["I1"] = "Enrollment"
    ws["J1"] = "Status"
    ws["K1"] = "Brief Summary"
    ws["L1"] = "Detailed Description"
    ws["M1"] = "Inclusion Criteria (may also contain the Exclusion Criteria due to parsing issues)"
    ws["N1"] = "Exclusion Criteria"
    ws["O1"] = "Conditions"
    ws["P1"] = "Primary Purpose"
    ws["Q1"] = "Primary Outcome Measures"
    ws["R1"] = "Secondary Outcome Measures"
    ws["S1"] = "Allocation"
    ws["T1"] = "Interventional Model"
    ws["U1"] = "Masking"
    ws["V1"] = "Study Start"
    ws["W1"] = "Primary Completion Date (Est)"
    ws["X1"] = "Locations"


    #sets up filtering for the excel sheet
    ws.auto_filter.ref = ws.dimensions
    #freezes first row so that the headers are always visible when scrolling through the data
    ws.freeze_panes = 'A2'

    #save and done
    wb.save('output.xlsx')
    print("Done")
