import urllib.request
import json
import ssl
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

col1 = "NCT"
col2 = "URL"
col3 = "Intervention/Treatment"
col4 = "Trial Name"
col5 = "Acronym"
col6 = "Sponsor"
col7 = "Funder Type"
col8 = "Phase"
col9 = "Enrollment"
col10 = "Status"
col11 = "Brief Summary"
col12 = "Detailed Description"
col13 = "Inclusion Criteria (may also contain the Exclusion Criteria due to parsing issues)"
col14 = "Exclusion Criteria"
col15 = "Conditions"
col16 = "Primary Purpose"
col17 = "Primary Outcome Measures"
col18 = "Secondary Outcome Measures"
col19 = "Allocation"
col20 = "Interventional Model"
col21 = "Masking"
col22 = "Study Start"
col23 = "Primary Completion Date (Est)"
col24 = "Locations"

def find_key_recursive(data, target_key):
    if isinstance(data, dict):
        for key, value in data.items():
            if key == target_key:
                return value
            result = find_key_recursive(value, target_key)
            if result is not None:
                return result
    elif isinstance(data, list):
        for item in data:
            result = find_key_recursive(item, target_key)
            if result is not None:
                return result
    return None

def locParse(dict):
    str2 = ""
    for key, value in dict.items():
        str = ""
        if (key != "facility" and key != "city" and key != "state" and key != "country" and key != "zip"):
            continue
        str += f", {value}"
        str2 += str
        if (len(str2)) > 1000:
            str2 += ", [TRUNCATED to 1000 characters for processing reasons]"
            break
    return str2[2:]

def splitCriteria(str):
    ind = str.find("Exclusion Criteria")
    if ind != -1:
        first_part = str[:ind]
        second_part = str[ind:]
        return [first_part, second_part]
    else:
        return [str]

ssl._create_default_https_context = ssl._create_unverified_context

ncts = []              #1
urls = []              #2
interventions = []     #3
names = []             #4
acronyms = []          #5
sponsors = []          #6
funderTypes = []       #7
phases = []            #8
enrolls = []           #9
statuses = []           #10
summaries = []          #11
descriptions = []       #12
criterias = []          #13
exclusions = []         #14
conditions = []         #15
primPurposes = []       #16
primOutcomes = []       #17
secOutcomes = []        #18
allocations = []        #19
interventionModels = []  #20
maskings = []            #21
starts = []              #22
primComps = []           #23
locs = []                #24

def mainLoop(nctNum):
    url: str = f'http://clinicaltrials.gov/api/v2/studies/{nctNum}'



    with urllib.request.urlopen(url) as response:
        if response.status != 200:
            print(f"Error: Received status code {response.status} for NCT {nctNum}")
            return
        j = response.read()
        data = json.loads(j)

        
    det_desc = find_key_recursive(data, "detailedDescription")
    #print("Detailed Description:", brief_summary)
    crit = find_key_recursive(data, "eligibilityCriteria")
        
    critList = splitCriteria(crit)
    name = find_key_recursive(data, "briefTitle")
    nct = find_key_recursive(data, "nctId")
    sponsor = find_key_recursive(data, "name")
    studyStart = find_key_recursive(data, "startDateStruct") 
    primComp = find_key_recursive(data, "primaryCompletionDateStruct") 
    enrollment = find_key_recursive(data, "enrollmentInfo") 
    phase = find_key_recursive(data, "phases") 
    brief_summary = find_key_recursive(data, "briefSummary")
    Intervention_treat = find_key_recursive(data, "interventionNames") #might have more than 1? 
    status = find_key_recursive(data, "overallStatus")
    designDetails = find_key_recursive(data, "designInfo") 
    loc = find_key_recursive(data, "locations")
    acronym = find_key_recursive(data, "acronym")
    condition = find_key_recursive(data, "conditions")
    funderType = find_key_recursive(data, "organization")["class"]
    outcomes = find_key_recursive(data, "outcomesModule")

    ncts.append(nct)
    urls.append(f"https://clinicaltrials.gov/study/{nctNum}")
    print(nctNum)
    if (Intervention_treat != None):
        interventions.append(' '.join(str(x) for x in Intervention_treat))
    else:
        x = find_key_recursive(data, "interventions")
        interventions.append(' '.join(str(i) for i in x))
    names.append(name)
    phases.append(' '.join(str(x) for x in phase))
    sponsors.append(sponsor)
    statuses.append(status)
    acronyms.append(acronym)
    summaries.append(brief_summary)
    descriptions.append(det_desc)

    criterias.append(critList[0])

    if (len(critList)) == 2:
        exclusions.append(critList[1])
    else:
        exclusions.append("")

    if (enrollment == None):
        enrolls.append("")
    else:
        enrolls.append(enrollment["count"])
    
    funderTypes.append(funderType)
    
    conditions.append(', '.join(str(x) for x in condition)[1:])
    
    if ("primaryPurpose" not in designDetails):
        primPurposes.append("")
    else:
        primPurposes.append(designDetails["primaryPurpose"])

    if (outcomes == None):
        primOutcomes.append("")
    else:
        primOutcomes.append(',\n'.join(str(x["measure"]) for x in outcomes["primaryOutcomes"]))
    
    if find_key_recursive(data, "secondaryOutcomes") is None:
        secOutcomes.append("")
    else:
        secOutcomes.append(',\n'.join(str(x["measure"]) for x in outcomes["secondaryOutcomes"]))
    
    if ("allocation" not in designDetails):
        allocations.append("")
    else:
        allocations.append(designDetails["allocation"])

    if ("interventionModel" not in designDetails):
        interventionModels.append("")
    else:
        interventionModels.append(designDetails["interventionModel"])

    if ("maskingInfo" not in designDetails):
        maskings.append("")
    else:
        maskings.append(designDetails["maskingInfo"]["masking"])
    
    if (studyStart == None):
        starts.append("")
    else:
        starts.append(studyStart["date"])

    if (primComp == None):
        primComps.append("")
    else:
        primComps.append(primComp["date"])
    
    if (loc == None):
        locs.append("")
    else:
        locs.append(',\n\n'.join(str(locParse(x)) for x in loc))
    print("done")



#########################################################
#Main
#########################################################

wb = load_workbook('input.xlsx')
ws = wb.active

column_list = [cell.value for cell in ws['A']]
column_list = column_list[1:]
for x in column_list:
    mainLoop(x)



df = pd.DataFrame({col1 : ncts, col2: urls, col3: interventions, col4: names, col5: acronyms, col6: sponsors, col7: funderTypes, col8: phases, col9: enrolls,  
                   col10: statuses, col11: summaries, col12: descriptions, col13: criterias, col14: exclusions,  col15: conditions, col16: primPurposes, 
                   col17: primOutcomes, col18: secOutcomes, col19: allocations, col20: interventionModels, col21: maskings, 
                   col22: starts, col23: primComps, col24: locs})

df.to_excel('test.xlsx', sheet_name='sheet1', index=False)

wb = load_workbook('test.xlsx')
ws = wb['sheet1']
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)

ws.column_dimensions['A'].width = 15 #1
ws.column_dimensions['B'].width = 25 #2
ws.column_dimensions['C'].width = 25 #3
ws.column_dimensions['D'].width = 25 #4
ws.column_dimensions['E'].width = 20 #5
ws.column_dimensions['F'].width = 25 #6
ws.column_dimensions['G'].width = 20 #7
ws.column_dimensions['H'].width = 15 #8
ws.column_dimensions['I'].width = 15 #9    
ws.column_dimensions['J'].width = 20 #10
ws.column_dimensions['K'].width = 50 #11
ws.column_dimensions['L'].width = 50 #12
ws.column_dimensions['M'].width = 65 #13
ws.column_dimensions['N'].width = 50 #14
ws.column_dimensions['O'].width = 25 #15
ws.column_dimensions['P'].width = 15 #16
ws.column_dimensions['Q'].width = 65 #17
ws.column_dimensions['R'].width = 65 #18
ws.column_dimensions['S'].width = 25 #19
ws.column_dimensions['T'].width = 20 #20
ws.column_dimensions['U'].width = 15 #21
ws.column_dimensions['V'].width = 15 #22
ws.column_dimensions['W'].width = 15 #23
ws.column_dimensions['X'].width = 35 #24

ws.auto_filter.ref = ws.dimensions

ws.freeze_panes = 'A2'

wb.save('test.xlsx')
